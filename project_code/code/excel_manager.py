import openpyxl
import os

class ExcelManager:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None
        self.sheet = None
        
    def load_workbook(self):
        """Carica il file Excel se esiste, altrimenti lo crea."""
        if os.path.exists(self.file_path):
            self.workbook = openpyxl.load_workbook(self.file_path)
            print(f"File {self.file_path} caricato con successo.")
        else:
            self.workbook = openpyxl.Workbook()
            self.workbook.save(self.file_path)
            print(f"File {self.file_path} creato con successo.")
        self.sheet = self.workbook.active
    
    def read_cell(self, cell):
        """Legge il valore di una cella specifica."""
        if self.sheet:
            return self.sheet[cell].value
        return None
    
    def write_cell(self, cell, value):
        """Scrive un valore in una cella e salva il file."""
        if self.sheet:
            self.sheet[cell] = value
            self.workbook.save(self.file_path)
            print(f"Valore '{value}' scritto in {cell}.")
    
    def close(self):
        """Chiude il file Excel."""
        if self.workbook:
            self.workbook.close()

    def process_all_files(data_folder):
        apartment_totals = {} 
        
        """Itera su tutte le cartelle e processa i file Excel."""
        for root, _, files in os.walk(data_folder):
            if "summary_apartment" in root:
                continue
            for file in files:
                if not file.endswith(".xlsx") or file.startswith("~$"):  # Ignora i file temporanei di Excel
                    continue
                if file.endswith(".xlsx"):
                    file_path = os.path.join(root, file)
                try:
                    excel = ExcelManager(file_path)
                    excel.load_workbook()
                    
                    apartment_name = excel.read_cell("C2")  # Nome dell'appartamento
                    c5 = excel.read_cell("C5") or 0  # Se None, mettiamo 0
                    c6 = excel.read_cell("C6") or 0  
                    total = c5 + c6
                    
                    excel.write_cell("C11", total)
                    apartment_totals[apartment_name] = total
                    
                    print(f"File {file}: {c5} + {c6} = {total} (scritto in C11)")
                    excel.close()
                except Exception as e:
                    print(f"Errore con il file {file}: {e}")

        return apartment_totals  # Ritorna SEMPRE un dizionario
                        
    def update_summary(apartment_totals):
        """Aggiorna il file summary.xlsx con i dati degli appartamenti."""
        base_dir = os.path.dirname(os.getcwd())  # Torna a 'mati_management_excel_apartment/project_code'

        # Costruisci il percorso corretto
        summary_path = os.path.join(base_dir, "data", "summary_apartment", "summary.xlsx")

        if not os.path.exists(summary_path):
            print(f"Errore: il file {summary_path} non esiste.")
            return

        try:
            summary_manager = ExcelManager(summary_path)
            summary_manager.load_workbook()

            sheet = summary_manager.sheet  # Ottieni il foglio di lavoro

            for row in range(3, sheet.max_row + 1):  # Dalla riga 3 in poi
                apartment_name = sheet[f"B{row}"].value  # Nome dell'appartamento
                if apartment_name in apartment_totals:
                    total = apartment_totals[apartment_name]
                    sheet[f"C{row}"] = total  # Scrive il valore nella colonna C

            summary_manager.workbook.save(summary_path)
            summary_manager.close()
            print("Summary.xlsx aggiornato con successo!")

        except Exception as e:
            print(f"Errore durante l'aggiornamento del summary.xlsx: {e}")